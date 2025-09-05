import os
from datetime import datetime, date, timedelta
from slack_sdk import WebClient
from openpyxl import load_workbook

SLACK_BOT_TOKEN = os.getenv("SLACK_BOT_TOKEN")
CHANNEL_IDS = os.getenv("SLACK_CHANNEL_IDS", "").split(",")

client = WebClient(token=SLACK_BOT_TOKEN)

def load_festivals():
    festivals = {}
    wb = load_workbook("holidays.xlsx")
    sheet = wb.active

    # Assume headers in first row: date | festival | location
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_date, festival, location = row
        if not raw_date:
            continue
        # Handle both Excel date cells and string like "06/09/2025"
        if isinstance(raw_date, datetime):
            holiday_date = raw_date.date()
        else:
            holiday_date = datetime.strptime(str(raw_date).strip(), "%d/%m/%Y").date()

        festivals[holiday_date] = {
            "festival": str(festival).strip(),
            "location": str(location).strip()
        }
    return festivals

def check_and_send():
    festivals = load_festivals()
    today = date.today()
    tomorrow = today + timedelta(days=1)
    print(f"📅 Today: {today}, Tomorrow: {tomorrow}")

    if tomorrow in festivals:
        details = festivals[tomorrow]
        message = (
            f"Tomorrow {tomorrow.strftime('%d-%B-%Y')} "
            f"on the occasion of {details['festival']} "
            f"is a holiday for {details['location']} location. 🎉"
        )
        for channel in CHANNEL_IDS:
            channel_id = channel.strip()
            if channel_id:
                try:
                    client.chat_postMessage(channel=channel_id, text=message)
                    print(f"✅ Sent to {channel_id}: {message}")
                except Exception as e:
                    print(f"❌ Failed to send to {channel_id}: {e}")
    else:
        print("ℹ️ No holiday tomorrow.")

if __name__ == "__main__":
    check_and_send()
